from io import BytesIO
from datetime import datetime, date
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Optional


MISSING_BIRTHDATE = date(1900, 1, 1)


class ReportGenerator:
    """Generate reports in PDF and Excel formats"""

    @staticmethod
    def _format_birthdate_and_age(user):
        if not user.birthdate or user.birthdate == MISSING_BIRTHDATE:
            return "", ""

        age = (datetime.now().date() - user.birthdate).days // 365
        return user.birthdate.strftime('%d.%m.%Y'), f"{age} р."

    @staticmethod
    def _build_report_rows(appointments: List, free_slots: Optional[List] = None):
        rows = [
            {
                "type": "appointment",
                "appointment": appointment,
                "start_time": appointment.start_time,
                "end_time": appointment.end_time
            }
            for appointment in appointments
        ]
        rows.extend(
            {
                "type": "free_slot",
                "slot": slot,
                "start_time": slot.start_time,
                "end_time": slot.end_time
            }
            for slot in (free_slots or [])
        )
        rows.sort(key=lambda row: (row["start_time"], 0 if row["type"] == "appointment" else 1))
        return rows

    @staticmethod
    def _format_day_separator(value: datetime) -> str:
        return value.strftime('%d.%m.%Y')

    @staticmethod
    def generate_pdf_report(
        appointments: List,
        from_date: str,
        to_date: str,
        free_slots: Optional[List] = None
    ) -> BytesIO:
        """Generate enhanced PDF report for appointments"""

        # Register fonts with Cyrillic support
        try:
            # Try Arial Unicode - has full Cyrillic support
            pdfmetrics.registerFont(TTFont('ArialUnicode', '/System/Library/Fonts/Supplemental/Arial Unicode.ttf'))
            pdfmetrics.registerFont(TTFont('Arial', '/System/Library/Fonts/Supplemental/Arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', '/System/Library/Fonts/Supplemental/Arial Bold.ttf'))
            font_name = 'ArialUnicode'
            font_name_bold = 'Arial-Bold'
        except Exception as e:
            print(f"Font registration error: {e}")
            # Try alternative paths
            try:
                pdfmetrics.registerFont(TTFont('ArialUnicode', '/Library/Fonts/Arial Unicode.ttf'))
                font_name = 'ArialUnicode'
                font_name_bold = 'ArialUnicode'
            except:
                # Last fallback - Helvetica (won't show Cyrillic correctly)
                font_name = 'Helvetica'
                font_name_bold = 'Helvetica-Bold'
                print("Warning: Using Helvetica - Cyrillic characters may not display correctly")

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=30
        )

        # Container for elements
        elements = []
        styles = getSampleStyleSheet()

        # Calculate statistics
        total_appointments = len(appointments)
        booked_count = sum(1 for a in appointments if a.status == 'booked')
        cancelled_count = sum(1 for a in appointments if a.status == 'cancelled')
        unique_patients = len(set(a.user_id for a in appointments))

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=20,
            textColor=colors.HexColor('#06b6d4'),
            spaceAfter=12,
            alignment=1,
            fontName=font_name_bold
        )

        title = Paragraph("Звіт про записи на прийом", title_style)
        elements.append(title)

        # Date range
        date_text = f"Період: {from_date} - {to_date}"
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#475569'),
            alignment=1,
            spaceAfter=20,
            fontName=font_name
        )
        elements.append(Paragraph(date_text, date_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Statistics summary
        stats_data = [
            ['Всього записів', 'Скасовано', 'Унікальних пацієнтів'],
            [str(total_appointments), str(cancelled_count), str(unique_patients)]
        ]

        stats_table = Table(stats_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06b6d4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e0f2fe')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#0c4a6e')),
            ('FONTNAME', (0, 1), (-1, 1), font_name_bold),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('TOPPADDING', (0, 1), (-1, 1), 12),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#06b6d4')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bae6fd'))
        ]))

        elements.append(stats_table)
        elements.append(Spacer(1, 0.4 * inch))

        # Section title
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=12,
            fontName=font_name_bold
        )
        elements.append(Paragraph("Детальна інформація про записи", section_style))
        elements.append(Spacer(1, 0.1 * inch))

        # Table data with more details
        data = [['№', 'Дата та час', 'Пацієнт', 'Телефон', 'Дата народж.', 'Вік', 'Нотатки']]

        current_day = None
        item_number = 0
        day_separator_rows = []

        for row_item in ReportGenerator._build_report_rows(appointments, free_slots):
            start_time = row_item["start_time"]
            row_day = start_time.date()
            if row_day != current_day:
                current_day = row_day
                day_separator_rows.append(len(data))
                data.append([ReportGenerator._format_day_separator(start_time), '', '', '', '', '', ''])

            item_number += 1
            if row_item["type"] == "free_slot":
                end_time = row_item["end_time"]
                row = [
                    str(item_number),
                    f"{start_time.strftime('%d.%m.%Y')}\n{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}",
                    "",
                    "",
                    "",
                    "",
                    ""
                ]
                data.append(row)
                continue

            appointment = row_item["appointment"]
            user = appointment.user
            birthdate_text, age_text = ReportGenerator._format_birthdate_and_age(user)

            row = [
                str(item_number),
                f"{appointment.start_time.strftime('%d.%m.%Y')}\n{appointment.start_time.strftime('%H:%M')}-{appointment.end_time.strftime('%H:%M')}",
                user.name,
                user.phone,
                birthdate_text,
                age_text,
                (appointment.notes or user.notes or '-')[:50]
            ]
            data.append(row)

        # Create table with adjusted widths
        table = Table(data, colWidths=[0.4*inch, 1.2*inch, 1.6*inch, 1.2*inch, 1.1*inch, 0.6*inch, 2*inch])

        # Enhanced table style
        table_style = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0891b2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),

            # Body styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]

        for row_idx in day_separator_rows:
            table_style.extend([
                ('SPAN', (0, row_idx), (-1, row_idx)),
                ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#e0f2fe')),
                ('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.HexColor('#0c4a6e')),
                ('FONTNAME', (0, row_idx), (-1, row_idx), font_name_bold),
                ('FONTSIZE', (0, row_idx), (-1, row_idx), 9),
                ('ALIGN', (0, row_idx), (-1, row_idx), 'LEFT'),
                ('TOPPADDING', (0, row_idx), (-1, row_idx), 7),
                ('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 7),
                ('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#06b6d4')),
                ('LINEBELOW', (0, row_idx), (-1, row_idx), 0.5, colors.HexColor('#bae6fd')),
            ])

        table.setStyle(TableStyle(table_style))
        elements.append(table)

        # Footer
        elements.append(Spacer(1, 0.3 * inch))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#64748b'),
            alignment=1,
            fontName=font_name
        )
        footer_text = f"Згенеровано: {datetime.now().strftime('%d.%m.%Y о %H:%M')} | Медичний центр - Система управління записами"
        elements.append(Paragraph(footer_text, footer_style))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_excel_report(
        appointments: List,
        from_date: str,
        to_date: str,
        free_slots: Optional[List] = None
    ) -> BytesIO:
        """Generate enhanced Excel report for appointments"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Записи на прийом"

        # Calculate statistics
        total_appointments = len(appointments)
        booked_count = sum(1 for a in appointments if a.status == 'booked')
        cancelled_count = sum(1 for a in appointments if a.status == 'cancelled')
        unique_patients = len(set(a.user_id for a in appointments))

        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "📊 Звіт про записи на прийом"
        title_cell.font = Font(size=18, bold=True, color="06b6d4")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Date range
        ws.merge_cells('A2:J2')
        date_cell = ws['A2']
        date_cell.value = f"Період: {from_date} - {to_date}"
        date_cell.font = Font(size=11, color="475569")
        date_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        # Statistics summary (row 4)
        stats_headers = ['Всього записів', 'Скасовано', 'Унікальних пацієнтів']
        stats_values = [total_appointments, cancelled_count, unique_patients]

        stats_fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
        stats_font = Font(bold=True, color="FFFFFF", size=11)
        value_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
        value_font = Font(bold=True, color="0c4a6e", size=14)

        for col_idx, (header, value) in enumerate(zip(stats_headers, stats_values), start=2):
            # Header
            header_cell = ws.cell(row=4, column=col_idx * 2 - 1)
            header_cell.value = header
            header_cell.fill = stats_fill
            header_cell.font = stats_font
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=4, start_column=col_idx * 2 - 1, end_row=4, end_column=col_idx * 2)

            # Value
            value_cell = ws.cell(row=5, column=col_idx * 2 - 1)
            value_cell.value = value
            value_cell.fill = value_fill
            value_cell.font = value_font
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=5, start_column=col_idx * 2 - 1, end_row=5, end_column=col_idx * 2)

        ws.row_dimensions[4].height = 25
        ws.row_dimensions[5].height = 30

        # Section title
        ws.merge_cells('A7:J7')
        section_cell = ws['A7']
        section_cell.value = "Детальна інформація про записи"
        section_cell.font = Font(size=12, bold=True, color="0f172a")
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[7].height = 20

        # Table headers (row 8)
        headers = ['№', 'Дата', 'Час', 'Пацієнт', 'Телефон', 'Дата народження', 'Вік', 'Нотатки (запис)', 'Нотатки (пацієнт)']
        header_fill = PatternFill(start_color="0891b2", end_color="0891b2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[8].height = 25

        # Data rows
        report_rows = ReportGenerator._build_report_rows(appointments, free_slots)
        row_idx = 9
        item_number = 0
        current_day = None
        max_report_col = 9

        for row_item in report_rows:
            start_time = row_item["start_time"]
            row_day = start_time.date()
            if row_day != current_day:
                current_day = row_day
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_report_col)
                separator_cell = ws.cell(row=row_idx, column=1)
                separator_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
                separator_cell.value = ReportGenerator._format_day_separator(start_time)
                separator_cell.fill = separator_fill
                separator_cell.font = Font(bold=True, color="0c4a6e", size=11)
                separator_cell.alignment = Alignment(horizontal='left', vertical='center')
                for col in range(1, max_report_col + 1):
                    ws.cell(row=row_idx, column=col).fill = separator_fill
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1

            item_number += 1
            if row_item["type"] == "free_slot":
                end_time = row_item["end_time"]
                row_data = [
                    item_number,
                    start_time.strftime('%d.%m.%Y'),
                    f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}",
                    "",
                    "",
                    "",
                    "",
                    "",
                    ""
                ]
            else:
                appointment = row_item["appointment"]
                user = appointment.user
                birthdate_text, age_text = ReportGenerator._format_birthdate_and_age(user)

                # Row data
                row_data = [
                    item_number,
                    appointment.start_time.strftime('%d.%m.%Y'),
                    f"{appointment.start_time.strftime('%H:%M')}-{appointment.end_time.strftime('%H:%M')}",
                    user.name,
                    user.phone,
                    birthdate_text,
                    age_text,
                    appointment.notes or '-',
                    user.notes or '-'
                ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left' if col_idx > 3 else 'center', vertical='center', wrap_text=True)
                cell.font = Font(size=10, color="1e293b")

            # Alternating row colors
            if item_number % 2 == 0:
                fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
                for col in range(1, max_report_col + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30

        # Footer
        last_row = row_idx + 1
        ws.merge_cells(f'A{last_row}:J{last_row}')
        footer_cell = ws[f'A{last_row}']
        footer_cell.value = f"Згенеровано: {datetime.now().strftime('%d.%m.%Y о %H:%M')} | Медичний центр - Система управління записами"
        footer_cell.font = Font(size=9, color="64748b", italic=True)
        footer_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[last_row].height = 18

        # Freeze panes (freeze header row)
        ws.freeze_panes = 'A9'

        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer
